"""
Parser robusto y flexible para archivos Excel de información exógena.
Maneja inconsistencias comunes en formato, estructura y datos.
"""
import pandas as pd
import openpyxl
from typing import Dict, List, Any, Tuple, Optional
from decimal import Decimal, InvalidOperation
import logging
import re
from datetime import datetime
import numpy as np

logger = logging.getLogger(__name__)


class RobustExogenaParser:
    """
    Parser robusto para archivos Excel de información exógena que maneja:
    - Posiciones variables de encabezados
    - Nombres de columnas inconsistentes
    - Formatos monetarios complejos
    - Metadatos intercalados
    - Validación robusta de datos
    """
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.stats = {
            'total_records': 0,
            'processed_records': 0,
            'skipped_records': 0,
            'total_income': Decimal('0'),
            'total_withholdings': Decimal('0'),
            'income_by_type': {},
            'income_by_schedule': {},
            'third_parties_count': 0,
            'concepts_count': 0
        }
        
        # Mapeo de aliases para columnas comunes
        self.column_aliases = {
            'nit': ['nit del tercero', 'nit_tercero', 'nit tercero', 'identificacion', 'documento', 'cedula'],
            'name': ['nombre del tercero', 'nombre_tercero', 'nombre tercero', 'razon social', 'tercero'],
            'concept': ['concepto', 'codigo concepto', 'codigo_concepto', 'cod concepto', 'tipo concepto'],
            'gross_amount': ['valor del pago o abono en cuenta', 'valor_pago', 'valor bruto', 'valor_bruto', 'monto', 'valor', 'importe'],
            'withholding': ['retencion practicada', 'retencion_practicada', 'valor retencion', 'retencion', 'ret_fuente']
        }
        
        # Clasificación de conceptos por código
        self.concept_mapping = {
            # Rentas de trabajo
            '5001': {'type': 'salary', 'schedule': 'labor', 'description': 'Salarios'},
            '5002': {'type': 'honorarios', 'schedule': 'labor', 'description': 'Honorarios'},
            '5003': {'type': 'services', 'schedule': 'labor', 'description': 'Servicios'},
            '5004': {'type': 'commissions', 'schedule': 'labor', 'description': 'Comisiones'},
            
            # Rentas de capital
            '5005': {'type': 'rental', 'schedule': 'capital', 'description': 'Arrendamientos'},
            '5006': {'type': 'interests', 'schedule': 'capital', 'description': 'Rendimientos financieros'},
            '5007': {'type': 'interests', 'schedule': 'capital', 'description': 'Intereses'},
            '5008': {'type': 'dividends', 'schedule': 'capital', 'description': 'Dividendos'},
            
            # Otros
            '5009': {'type': 'other', 'schedule': 'other', 'description': 'Otros ingresos'},
            '5010': {'type': 'prizes', 'schedule': 'other', 'description': 'Premios y rifas'}
        }
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Método principal para parsear archivos Excel de información exógena.
        """
        try:
            logger.info(f"Iniciando procesamiento robusto de: {file_path}")
            
            # Resetear estadísticas y errores
            self._reset_stats()
            
            # Detectar formato del archivo
            file_info = self._analyze_file_structure(file_path)
            if not file_info['valid']:
                return self._error_response(file_info['errors'])
            
            # Cargar y procesar datos
            raw_data = self._load_excel_data(file_path, file_info)
            if raw_data.empty:
                return self._error_response(['No se encontraron datos válidos en el archivo'])
            
            # Detectar y mapear columnas
            column_mapping = self._detect_and_map_columns(raw_data)
            if not column_mapping:
                return self._error_response(['No se pudieron identificar las columnas requeridas'])
            
            # Limpiar y validar datos
            clean_data = self._clean_and_validate_data(raw_data, column_mapping)
            
            # Procesar registros
            processed_records = self._process_records(clean_data, column_mapping)
            
            # Calcular estadísticas
            self._calculate_statistics(processed_records)
            
            # Validar consistencia de datos
            validation_result = self._validate_data_consistency(processed_records)
            self.warnings.extend(validation_result['warnings'])
            
            logger.info(f"Procesamiento completado: {len(processed_records)} registros")
            
            return {
                'success': True,
                'records': processed_records,
                'stats': self.stats,
                'file_info': file_info,
                'column_mapping': column_mapping,
                'validation': validation_result,
                'errors': self.errors,
                'warnings': self.warnings
            }
            
        except Exception as e:
            logger.error(f"Error crítico en parser: {str(e)}")
            return self._error_response([f"Error crítico: {str(e)}"])
    
    def _analyze_file_structure(self, file_path: str) -> Dict[str, Any]:
        """Analiza la estructura del archivo Excel para detectar formato y problemas."""
        try:
            # Detectar tipo de archivo
            file_extension = file_path.lower().split('.')[-1]
            if file_extension not in ['xlsx', 'xls', 'csv']:
                return {'valid': False, 'errors': [f'Formato de archivo no soportado: {file_extension}']}
            
            # Análisis básico con openpyxl para Excel
            if file_extension in ['xlsx', 'xls']:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                sheet_names = workbook.sheetnames
                
                # Analizar primera hoja activa
                sheet = workbook.active
                total_rows = sheet.max_row
                total_cols = sheet.max_column
                
                # Detectar posible ubicación de encabezados
                header_row = self._detect_header_row(sheet)
                
                workbook.close()
                
                return {
                    'valid': True,
                    'file_type': file_extension,
                    'sheet_names': sheet_names,
                    'total_rows': total_rows,
                    'total_cols': total_cols,
                    'header_row': header_row,
                    'errors': []
                }
            else:
                # Análisis para CSV
                return {
                    'valid': True,
                    'file_type': 'csv',
                    'sheet_names': ['csv'],
                    'header_row': 0,
                    'errors': []
                }
                
        except Exception as e:
            return {'valid': False, 'errors': [f'Error analizando estructura: {str(e)}']}
    
    def _detect_header_row(self, sheet) -> int:
        """Detecta la fila que contiene los encabezados de columnas."""
        # Buscar en las primeras 15 filas
        for row_idx in range(1, min(16, sheet.max_row + 1)):
            row_values = [cell.value for cell in sheet[row_idx] if cell.value]
            
            # Buscar palabras clave típicas de encabezados
            header_keywords = ['nit', 'tercero', 'concepto', 'valor', 'retencion', 'pago', 'abono']
            
            if len(row_values) >= 3:  # Al menos 3 columnas con datos
                text_values = [str(val).lower() for val in row_values if val]
                keyword_matches = sum(1 for keyword in header_keywords 
                                    if any(keyword in text for text in text_values))
                
                if keyword_matches >= 2:  # Al menos 2 keywords encontradas
                    return row_idx - 1  # Retornar índice base 0 para pandas
        
        return 0  # Default a primera fila
    
    def _load_excel_data(self, file_path: str, file_info: Dict) -> pd.DataFrame:
        """Carga los datos del archivo Excel usando la información de estructura detectada."""
        try:
            if file_info['file_type'] == 'csv':
                # Intentar diferentes encodings para CSV
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding, 
                                       skiprows=file_info['header_row'])
                        return df
                    except UnicodeDecodeError:
                        continue
                raise Exception("No se pudo decodificar el archivo CSV con ninguna codificación")
            
            else:
                # Cargar Excel
                df = pd.read_excel(file_path, 
                                 sheet_name=0,  # Primera hoja
                                 header=file_info['header_row'],
                                 engine='openpyxl')
                return df
                
        except Exception as e:
            logger.error(f"Error cargando datos: {str(e)}")
            return pd.DataFrame()
    
    def _detect_and_map_columns(self, df: pd.DataFrame) -> Dict[str, str]:
        """Detecta y mapea las columnas del DataFrame a campos estándar."""
        column_mapping = {}
        df_columns_lower = [col.lower().strip() for col in df.columns]
        
        for standard_field, aliases in self.column_aliases.items():
            matched = False
            
            for alias in aliases:
                for idx, col_name in enumerate(df_columns_lower):
                    if alias in col_name or col_name in alias:
                        column_mapping[standard_field] = df.columns[idx]
                        matched = True
                        break
                if matched:
                    break
            
            if not matched:
                # Buscar coincidencias parciales más flexibles
                for idx, col_name in enumerate(df_columns_lower):
                    if standard_field == 'nit' and any(keyword in col_name for keyword in ['nit', 'ident', 'doc']):
                        column_mapping[standard_field] = df.columns[idx]
                        break
                    elif standard_field == 'name' and any(keyword in col_name for keyword in ['nombre', 'tercero', 'razon']):
                        column_mapping[standard_field] = df.columns[idx]
                        break
                    elif standard_field == 'concept' and 'concepto' in col_name:
                        column_mapping[standard_field] = df.columns[idx]
                        break
                    elif standard_field == 'gross_amount' and any(keyword in col_name for keyword in ['valor', 'pago', 'monto']):
                        column_mapping[standard_field] = df.columns[idx]
                        break
                    elif standard_field == 'withholding' and 'retencion' in col_name:
                        column_mapping[standard_field] = df.columns[idx]
                        break
        
        # Verificar que se encontraron las columnas mínimas requeridas
        required_fields = ['nit', 'name', 'gross_amount']
        missing_fields = [field for field in required_fields if field not in column_mapping]
        
        if missing_fields:
            self.warnings.append(f"Campos requeridos no encontrados: {missing_fields}")
            
        logger.info(f"Mapeo de columnas detectado: {column_mapping}")
        return column_mapping
    
    def _clean_and_validate_data(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> pd.DataFrame:
        """Limpia y valida los datos del DataFrame."""
        # Crear copia para trabajar
        clean_df = df.copy()
        
        # Eliminar filas completamente vacías
        clean_df = clean_df.dropna(how='all')
        
        # Filtrar filas que parecen ser metadatos (filas con muy pocos valores o texto descriptivo)
        if len(clean_df) > 0:
            # Eliminar filas donde la columna de NIT tiene valores no numéricos obvios
            if 'nit' in column_mapping:
                nit_col = column_mapping['nit']
                mask = clean_df[nit_col].astype(str).str.contains(r'\d', na=False)
                clean_df = clean_df[mask]
        
        # Aplicar limpieza específica por campo
        for standard_field, original_col in column_mapping.items():
            if original_col in clean_df.columns:
                if standard_field == 'nit':
                    clean_df[original_col] = clean_df[original_col].apply(self._clean_nit)
                elif standard_field in ['gross_amount', 'withholding']:
                    clean_df[original_col] = clean_df[original_col].apply(self._clean_amount)
                elif standard_field == 'concept':
                    clean_df[original_col] = clean_df[original_col].apply(self._clean_concept)
                elif standard_field == 'name':
                    clean_df[original_col] = clean_df[original_col].apply(self._clean_name)
        
        return clean_df
    
    def _process_records(self, df: pd.DataFrame, column_mapping: Dict[str, str]) -> List[Dict[str, Any]]:
        """Procesa los registros limpios y los convierte al formato estándar."""
        processed_records = []
        
        for idx, row in df.iterrows():
            try:
                # Extraer datos básicos
                record = {
                    'source_row': idx + 1,
                    'third_party_nit': row.get(column_mapping.get('nit', ''), ''),
                    'third_party_name': row.get(column_mapping.get('name', ''), ''),
                    'concept_code': self._extract_concept_code(row.get(column_mapping.get('concept', ''), '')),
                    'gross_amount': row.get(column_mapping.get('gross_amount', ''), Decimal('0')),
                    'withholding_amount': row.get(column_mapping.get('withholding', ''), Decimal('0'))
                }
                
                # Validar registro básico
                if not record['third_party_nit'] or record['gross_amount'] == Decimal('0'):
                    self.stats['skipped_records'] += 1
                    continue
                
                # Clasificar ingreso
                classification = self._classify_income(record['concept_code'], record['third_party_name'])
                record.update(classification)
                
                # Agregar metadatos adicionales
                record['concept_description'] = self.concept_mapping.get(
                    record['concept_code'], {}
                ).get('description', 'Concepto no clasificado')
                
                processed_records.append(record)
                self.stats['processed_records'] += 1
                
            except Exception as e:
                logger.warning(f"Error procesando fila {idx}: {str(e)}")
                self.warnings.append(f"Fila {idx} omitida por error: {str(e)}")
                self.stats['skipped_records'] += 1
                continue
        
        return processed_records
    
    def _clean_nit(self, nit_value) -> str:
        """Limpia y normaliza valores de NIT."""
        if pd.isna(nit_value) or nit_value == '':
            return ''
        
        # Convertir a string y limpiar
        nit_str = str(nit_value).strip()
        
        # Remover prefijos comunes
        nit_str = re.sub(r'^(NIT|CC|CE)[\s\-\.]*', '', nit_str, flags=re.IGNORECASE)
        
        # Remover caracteres no numéricos excepto guión
        nit_str = re.sub(r'[^\d\-]', '', nit_str)
        
        # Remover dígito de verificación (después del guión)
        nit_str = nit_str.split('-')[0]
        
        return nit_str
    
    def _clean_amount(self, amount_value) -> Decimal:
        """Limpia y convierte valores monetarios a Decimal."""
        if pd.isna(amount_value) or amount_value == '':
            return Decimal('0')
        
        try:
            # Convertir a string
            amount_str = str(amount_value).strip()
            
            # Remover símbolos de moneda y espacios
            amount_str = re.sub(r'[$\s]', '', amount_str)
            
            # Manejar formato colombiano (punto como separador de miles, coma como decimal)
            if ',' in amount_str and '.' in amount_str:
                # Formato: 1.234.567,89
                if amount_str.rfind(',') > amount_str.rfind('.'):
                    amount_str = amount_str.replace('.', '').replace(',', '.')
                # Formato: 1,234,567.89
                else:
                    amount_str = amount_str.replace(',', '')
            elif ',' in amount_str:
                # Solo coma - verificar si es decimal o separador de miles
                parts = amount_str.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    # Es decimal: 1234,56
                    amount_str = amount_str.replace(',', '.')
                else:
                    # Es separador de miles: 1,234,567
                    amount_str = amount_str.replace(',', '')
            
            # Remover puntos como separadores de miles si quedan
            if '.' in amount_str:
                parts = amount_str.split('.')
                if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) > 2):
                    # Múltiples puntos o parte decimal muy larga = separadores de miles
                    amount_str = ''.join(parts[:-1]) + '.' + parts[-1] if len(parts[-1]) <= 2 else ''.join(parts)
            
            return Decimal(amount_str)
            
        except (InvalidOperation, ValueError):
            logger.warning(f"No se pudo convertir amount: {amount_value}")
            return Decimal('0')
    
    def _clean_concept(self, concept_value) -> str:
        """Limpia valores de concepto."""
        if pd.isna(concept_value):
            return ''
        
        concept_str = str(concept_value).strip()
        
        # Extraer código numérico si está presente
        match = re.search(r'\b\d{4}\b', concept_str)
        if match:
            return match.group()
        
        return concept_str
    
    def _clean_name(self, name_value) -> str:
        """Limpia nombres de terceros."""
        if pd.isna(name_value):
            return ''
        
        name_str = str(name_value).strip().upper()
        
        # Remover caracteres especiales excesivos
        name_str = re.sub(r'[^\w\s\.\-&]', '', name_str)
        
        # Normalizar espacios
        name_str = re.sub(r'\s+', ' ', name_str)
        
        return name_str
    
    def _extract_concept_code(self, concept_value) -> str:
        """Extrae el código de concepto de diferentes formatos."""
        if pd.isna(concept_value):
            return ''
        
        concept_str = str(concept_value).strip()
        
        # Buscar código de 4 dígitos
        match = re.search(r'\b\d{4}\b', concept_str)
        if match:
            return match.group()
        
        # Buscar código al inicio seguido de guión o espacio
        match = re.search(r'^(\d+)[\s\-]', concept_str)
        if match:
            return match.group(1)
        
        return concept_str
    
    def _classify_income(self, concept_code: str, third_party_name: str) -> Dict[str, str]:
        """Clasifica el tipo de ingreso basado en el concepto y nombre del tercero."""
        # Primero intentar por código de concepto
        if concept_code in self.concept_mapping:
            mapping = self.concept_mapping[concept_code]
            return {
                'income_type': mapping['type'],
                'tax_schedule': mapping['schedule']
            }
        
        # Clasificación por patrones en nombre del tercero
        name_lower = third_party_name.lower()
        
        # Patrones para diferentes tipos de ingreso
        if any(keyword in name_lower for keyword in ['banco', 'financiera', 'cooperativa', 'fondo']):
            return {'income_type': 'interests', 'tax_schedule': 'capital'}
        
        elif any(keyword in name_lower for keyword in ['inmobiliaria', 'arrendamiento', 'propiedad']):
            return {'income_type': 'rental', 'tax_schedule': 'capital'}
        
        elif any(keyword in name_lower for keyword in ['consultoria', 'profesional', 'servicios']):
            return {'income_type': 'honorarios', 'tax_schedule': 'labor'}
        
        elif any(keyword in name_lower for keyword in ['empresa', 'compañia', 'corporacion', 'sas', 'sa', 'ltda']):
            return {'income_type': 'salary', 'tax_schedule': 'labor'}
        
        # Default
        return {'income_type': 'other', 'tax_schedule': 'other'}
    
    def _calculate_statistics(self, records: List[Dict[str, Any]]):
        """Calcula estadísticas completas de los registros procesados."""
        self.stats['total_records'] = len(records)
        
        # Totales
        self.stats['total_income'] = sum(record['gross_amount'] for record in records)
        self.stats['total_withholdings'] = sum(record['withholding_amount'] for record in records)
        
        # Estadísticas por tipo de ingreso
        income_by_type = {}
        for record in records:
            income_type = record['income_type']
            if income_type not in income_by_type:
                income_by_type[income_type] = {
                    'count': 0,
                    'gross_amount': Decimal('0'),
                    'withholding_amount': Decimal('0')
                }
            
            income_by_type[income_type]['count'] += 1
            income_by_type[income_type]['gross_amount'] += record['gross_amount']
            income_by_type[income_type]['withholding_amount'] += record['withholding_amount']
        
        self.stats['income_by_type'] = income_by_type
        
        # Estadísticas por cédula tributaria
        income_by_schedule = {}
        for record in records:
            schedule = record['tax_schedule']
            if schedule not in income_by_schedule:
                income_by_schedule[schedule] = {
                    'count': 0,
                    'gross_amount': Decimal('0'),
                    'withholding_amount': Decimal('0')
                }
            
            income_by_schedule[schedule]['count'] += 1
            income_by_schedule[schedule]['gross_amount'] += record['gross_amount']
            income_by_schedule[schedule]['withholding_amount'] += record['withholding_amount']
        
        self.stats['income_by_schedule'] = income_by_schedule
        
        # Contadores únicos
        self.stats['third_parties_count'] = len(set(record['third_party_nit'] for record in records))
        self.stats['concepts_count'] = len(set(record['concept_code'] for record in records))
    
    def _validate_data_consistency(self, records: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Valida la consistencia de los datos procesados."""
        validation_warnings = []
        validation_errors = []
        
        # Validar NITs duplicados con montos diferentes
        nit_groups = {}
        for record in records:
            nit = record['third_party_nit']
            if nit in nit_groups:
                if nit_groups[nit]['name'] != record['third_party_name']:
                    validation_warnings.append(f"NIT {nit} tiene nombres diferentes: '{nit_groups[nit]['name']}' vs '{record['third_party_name']}'")
            else:
                nit_groups[nit] = {
                    'name': record['third_party_name'],
                    'total_amount': record['gross_amount']
                }
        
        # Validar retenciones vs ingresos
        for record in records:
            if record['withholding_amount'] > record['gross_amount']:
                validation_warnings.append(f"Retención mayor que ingreso bruto en fila {record['source_row']}")
        
        # Validar códigos de concepto
        unknown_concepts = []
        for record in records:
            if record['concept_code'] not in self.concept_mapping and record['concept_code'] not in unknown_concepts:
                unknown_concepts.append(record['concept_code'])
        
        if unknown_concepts:
            validation_warnings.append(f"Códigos de concepto no reconocidos: {unknown_concepts}")
        
        return {
            'valid': len(validation_errors) == 0,
            'warnings': validation_warnings,
            'errors': validation_errors
        }
    
    def _reset_stats(self):
        """Resetea estadísticas para nuevo procesamiento."""
        self.errors = []
        self.warnings = []
        self.stats = {
            'total_records': 0,
            'processed_records': 0,
            'skipped_records': 0,
            'total_income': Decimal('0'),
            'total_withholdings': Decimal('0'),
            'income_by_type': {},
            'income_by_schedule': {},
            'third_parties_count': 0,
            'concepts_count': 0
        }
    
    def _error_response(self, errors: List[str]) -> Dict[str, Any]:
        """Genera respuesta de error estándar."""
        return {
            'success': False,
            'records': [],
            'stats': self.stats,
            'file_info': {},
            'column_mapping': {},
            'validation': {'valid': False, 'warnings': [], 'errors': errors},
            'errors': errors,
            'warnings': self.warnings
        }


# Alias para mantener compatibilidad con el código existente
ExogenaParser = RobustExogenaParser
